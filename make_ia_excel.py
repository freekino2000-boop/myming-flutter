import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 공통 스타일 ────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_all(thin=True):
    s = Side(style="thin" if thin else "medium", color="C8DCFF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_group():
    s = Side(style="medium", color="2C7FFF")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, row, col, value,
               bg="FFFFFF", fg="1A1A1A", bold=False,
               wrap=True, align_h="left", align_v="center",
               size=9, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = Font(name="맑은 고딕", size=size, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align_h, vertical=align_v,
                             wrap_text=wrap)
    if border:
        c.border = border_all()
    return c

def header_row(ws, row, values, col_start=1):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=v)
        c.fill = fill("0D1A3E")
        c.font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = border_all()

def group_row(ws, row, label, ncols, bg="E8F0FF", fg="2C7FFF"):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = fill(bg)
    c.font = Font(name="맑은 고딕", size=9, bold=True, color=fg)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=False)
    c.border = border_group()


# ════════════════════════════════════════════════
# 시트 1 — 화면 IA 전체표
# ════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "01_화면IA"

ws1.row_dimensions[1].height = 14   # 타이틀 행
ws1.row_dimensions[2].height = 30   # 헤더 행

# 타이틀
ws1.merge_cells("A1:J1")
t = ws1["A1"]
t.value = "마이밍 Flutter — 화면 IA 전체표"
t.fill = fill("1A3A7A")
t.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
t.alignment = Alignment(horizontal="left", vertical="center")

# 헤더
headers = ["#", "화면명", "라우트", "진입 경로", "탭 위치",
           "주요 컴포넌트", "사용 모델", "AppState 필드",
           "아웃고잉 네비게이션", "파일 경로"]
header_row(ws1, 2, headers)

# 열 너비
col_widths = [6, 18, 18, 22, 10, 38, 22, 28, 30, 38]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

r = 3  # 현재 행

# 데이터 정의: (그룹레이블, 행들)
# 행: (#, 화면명, 라우트, 진입, 탭, 컴포넌트, 모델, 상태, 아웃, 파일)
NCOLS = 10

sections = [
    ("🔐 인증 (Auth)", [
        ("A-01", "LoginPage\n로그인 화면", "/login", "initialRoute", "—",
         "소셜로그인 버튼 ×4\n_SocialBtn\n로고 영역\n약관 텍스트",
         "—", "—",
         "→ / (replace)\n로그인 성공",
         "pages/auth/login_page.dart"),
    ]),
    ("🏠 메인 셸 (IndexedStack)", [
        ("S-01", "MainShell\n하단 탭 컨테이너", "/", "replace (로그인 후)", "Root",
         "IndexedStack\nBottomNavigationBar\n_NavItem ×3",
         "—", "—",
         "Tab 0 → HomePage\nTab 1 → CardNewsPage\nTab 2 → GameSelectPage",
         "pages/main_shell.dart"),
    ]),
    ("🏠 홈 탭 (Tab 0)", [
        ("H-01", "HomePage\n만보기 메인 허브", "/ (tab 0)", "MainShell Tab 0", "Tab 0",
         "SliverAppBar\nPedometerCard\nBillHistoryEntry\nQuickActionSection\n_MissionSummaryCard\n_QuickMenuGrid (4���)",
         "—",
         "steps\nwalletAmount\ncanAttendToday\ncnMissionDone",
         "→ /profile\n→ /shop\n→ /offerwall\n→ /bill\n→ /fortune\n→ /coupon",
         "pages/home/home_page.dart"),
        ("H-02", "PedometerCard\n만보기 위젯", "—", "H-01 내 위젯", "—",
         "진행바 (LinearProgressIndicator)\n걸음 수 표시\n적립금 표시",
         "—", "steps\nwalletAmount", "—",
         "home/widgets/pedometer_card.dart"),
        ("H-03", "QuickActionSection\n출석·광고·행운", "—", "H-01 내 위젯", "—",
         "출석체크 버튼 (달력 팝업 포함)\n광고 보너스 (→ 카드뉴스 탭 이동)\n행운 뽑기 (10~100원 랜덤)",
         "—", "canAttendToday\nattendanceDates\nearn()\nswitchToCardNewsTab()", "→ CardNewsPage (탭 전환)",
         "home/widgets/quick_action_card.dart"),
        ("H-04", "BillHistoryEntry\n절감머니 내역 진입", "—", "H-01 내 위젯", "—",
         "절감 금액 표시\n→ /bill 진입 버튼",
         "—", "walletAmount", "→ /bill",
         "home/widgets/bill_history_entry.dart"),
    ]),
    ("📰 카드뉴스 탭 (Tab 1)", [
        ("N-01", "CardNewsPage\n카드뉴스 목록", "/ (tab 1)", "MainShell Tab 1", "Tab 1",
         "_CatChip ×5 (카테고리 필터)\n_SectionHeader\n_Top3Card (일간 TOP3)\n_WeeklyCard (주간, 카드 전체 클릭 가능)\n_ReactBtn",
         "CardNews\nNewsCategory",
         "cnReadCount\ncnMissionDone\nreadCNArticle()",
         "→ ArticleViewPage (push)\n→ /shop",
         "pages/card_news/card_news_page.dart"),
        ("N-02", "ArticleViewPage\n기사 상세", "—", "push (N-01 카드 탭)", "—",
         "카드 슬라이드 뷰어\n카테고리 배지\n반응 버튼\n_EarnBar (광고 버튼)\n광고 배너 (TODO)",
         "CardNews",
         "earn() 직접 호출\n(TODO: 광고 SDK 콜백으로 교체)",
         "← pop",
         "pages/card_news/article_view_page.dart"),
    ]),
    ("📰 카드뉴스 · SPA 전용", [
        ("N-03", "adVideoListPage\n광고보고 기사읽기", "/adVideoList", "H-03 (광고 보너스)", "—",
         "미션 배너 (읽은 기사 카운터)\nCARD_NEWS 기반 기사 그리드\nwatchAdAndOpenArticle(id)\nearnCN(id)",
         "CARD_NEWS",
         "STATE.reads\nplayFullscreenAd()\nopenCNArticle()",
         "SPA only — Flutter 미구현",
         "마이밍_만보기버전.html (adVideoListPage)"),
    ]),
    ("🎮 게임 탭 (Tab 2)", [
        ("G-01", "GameSelectPage\n게임 선택 2×4 그리드", "/game-select", "MainShell Tab 2", "Tab 2",
         "2×4 GridView\n게임 카드 ×8\nRANK 뱃지\n개발예정 스낵바",
         "—", "walletAmount",
         "→ BrickGamePage\n→ JumpGamePage\n→ CatchGamePage\n→ SnakeGamePage\n→ /rank\n(개발예정: puzzle·quiz·shooting·card)",
         "pages/game/game_select_page.dart"),
        ("G-02", "BrickGamePage\n벽돌깨기 (Flame)", "—", "push (G-01)", "—",
         "GameWidget<BrickGame>\n_ResultOverlay\nHUD (점수·레벨)\n패들 드래그",
         "—", "earn()",
         "← pop\n보상: min(score÷10, 200)",
         "game/games/brick_game_page.dart"),
        ("G-03", "JumpGamePage\n무한점프 (Flame)", "—", "push (G-01)", "—",
         "GameWidget<JumpGame>\nTapCallbacks\n장애물 스폰\n코인 수집",
         "—", "earn()",
         "← pop\n보상: min(dist÷20+coins×5, 150)",
         "game/games/jump_game_page.dart"),
        ("G-04", "CatchGamePage\n낚시대전 (Flame)", "—", "push (G-01)", "—",
         "GameWidget<CatchGame>\n드래그 조작\n🐟+10 🪙+15 💣-10",
         "—", "earn()",
         "← pop\n보상: min(score÷15, 100)",
         "game/games/catch_game_page.dart"),
        ("G-05", "SnakeGamePage\n뱀게임 (CustomPainter)", "—", "push (G-01)", "—",
         "Stack → Column + Positioned.fill\nCustomPaint (_SnakePainter)\n방향키 패드 (_DirBtn×4)\n스와이프 GestureDetector",
         "—", "earn()",
         "← pop\n보상: min(score×3, 150)",
         "game/games/snake_game_page.dart"),
        ("G-06", "GlobalRankPage\n전체 랭킹", "/rank", "push (G-01)", "—",
         "ListView (_RankEntry ×10)\nTOP3 강조 (👑🥈🥉)",
         "—", "—",
         "← pop",
         "pages/game/global_rank_page.dart"),
    ]),
    ("🎁 교환소 (Shop)", [
        ("SP-01", "ShopPage\n적립금 교환소", "/shop", "push (H-01/N-01)", "—",
         "_DwellTimeCard (체류 시간 카드)\n잔액 표시 배너\nTabBar ×8\n(부스터/이용권/상품권/주거/식음료/뷰티/디지털/기타)\n_ShopItemList\n_ShopItemCard ×19",
         "—",
         "walletAmount\nbuyBooster()\nspend()\nhasManhwaBoost\nfirstAccessDate\ntotalSecondsWithCurrent",
         "← pop",
         "pages/shop/shop_page.dart"),
    ]),
    ("🎯 고정비 절감 미션 (Offerwall)", [
        ("OW-01", "OfferwallPage\n29개 절감 미션", "/offerwall", "push (H-01/H-04)", "—",
         "_CatChip ×9 (카테고리)\n_TypeTab ×4 (전체/추천/간단/고보상)\n_OfferCard\n_OfferModal (3단계)\n배너 그라디언트",
         "Offer\nOfferCategory\nOfferType",
         "walletAmount\nearn()",
         "BottomSheet → _OfferModal\n← pop",
         "pages/offerwall/offerwall_page.dart"),
    ]),
    ("📋 고정비 관리 (Bill)", [
        ("B-01", "BillPage\n청구서 목록", "/bill", "push (H-01/H-04)", "—",
         "월간 합계 헤더\n절감 진행바\n탭 ×3 (전체/이번달/납부내역)\n_BillCard ×n\nFAB (추가)\n_LedgerTab (탭index==2)\n_BillFormSheet",
         "Bill\nBillType",
         "walletAmount\nspend()\nledger",
         "BottomSheet → _BillFormSheet\n탭 전환 → _LedgerTab\n← pop",
         "pages/bill/bill_page.dart"),
    ]),
    ("🔮 오늘의 운세 (Fortune)", [
        ("F-01", "FortunePage\n운세 조회", "/fortune", "push (H-01)", "—",
         "_FormView (생년월일·성별·시)\n_ResultView\nAnimationController (카드 뒤집기)\n_FortuneStat ×4",
         "—",
         "earn() (+₩5)",
         "← pop",
         "pages/fortune/fortune_page.dart"),
    ]),
    ("🎟 쿠폰함 (Coupon)", [
        ("C-01", "CouponPage\n보유 쿠폰 목록", "/coupon", "push (H-01)", "—",
         "쿠폰 카드 목록\nD-Day 뱃지\n코드 복사 버튼\nClipboard.setData()",
         "—", "—",
         "← pop",
         "pages/coupon/coupon_page.dart"),
    ]),
    ("👤 내 정보 (Profile)", [
        ("P-01", "ProfilePage\n사용자 정보·내역", "/profile", "push (H-01 앱바)", "—",
         "프로필 편집 영역\n_StatTile ×2 (적립금/절감)\n최근 내역 5건 (earn/spend 구분)\n_SettingItem 목록",
         "—",
         "walletAmount\nledger\nsteps",
         "→ /shop\n← pop",
         "pages/profile/profile_page.dart"),
    ]),
]

ROW_BG_ALT  = "F8FAFF"
ROW_BG_NORM = "FFFFFF"
even = False

for group_label, rows in sections:
    # 그룹 행
    group_row(ws1, r, group_label, NCOLS)
    ws1.row_dimensions[r].height = 16
    r += 1

    for row_data in rows:
        even = not even
        bg = ROW_BG_ALT if even else ROW_BG_NORM
        for col, val in enumerate(row_data, 1):
            b = True
            align = "left"
            sz = 9
            bold = False
            fg = "1A1A1A"
            row_bg = bg

            if col == 1:   # # 번호
                align = "center"; bold = True; fg = "2C7FFF"; row_bg = "EFF5FF"
            elif col == 3: # 라우트
                fg = "1E3A6E"; row_bg = "F0F4FF"
            elif col == 6: # 컴포넌트 (넓음)
                fg = "374151"
            elif col == 9: # 아웃고잉
                fg = "00875A"
            elif col == 10: # 파일
                fg = "6B7280"; sz = 8

            cell_style(ws1, r, col, val,
                       bg=row_bg, fg=fg, bold=bold,
                       wrap=True, align_h=align,
                       size=sz)
        ws1.row_dimensions[r].height = max(
            15, len(str(row_data[5]).split("\n")) * 13
        )
        r += 1

# 틀 고정
ws1.freeze_panes = "A3"


# ════════════════════════════════════════════════
# 시트 2 — 데이터 모델
# ════════════════════════════════════════════════
ws2 = wb.create_sheet("02_데이터모델")

ws2.merge_cells("A1:F1")
t2 = ws2["A1"]
t2.value = "마이밍 Flutter — 데이터 모델"
t2.fill = fill("1A3A7A")
t2.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
t2.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 14

headers2 = ["모델", "파일", "주요 필드", "Enum 정의", "목(Mock) 데이터", "사용 화면"]
header_row(ws2, 2, headers2)
ws2.row_dimensions[2].height = 28

col_w2 = [18, 34, 40, 42, 22, 20]
for i, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

models = [
    ("AppState\n(ChangeNotifier)",
     "models/app_state.dart",
     "int steps\ndouble walletAmount\nbool hasManhwaBoost\nbool hasSpeed2x\nbool hasSpeed5x\nbool hasAutoCollect\nint cnReadCount\nbool cnMissionDone\nList<LedgerEntry> ledger\nString? lastAttendanceDate\nList<String> attendanceDates (달력용)\nint cardNewsTabTrigger\nString? firstAccessDate\nint totalSecondsInApp\nDateTime? _sessionStart (runtime)",
     "—",
     "—",
     "전체 화면 (Provider)"),
    ("CardNews\n(immutable)",
     "models/card_news_model.dart",
     "String id, headline, summary, content\nNewsCategory category\nNewsTier tier\nString emoji, stat, time\nList<String> keywords\nint likes, bookmarks\nbool isNew, isRead\nbool isLiked, isBookmarked",
     "NewsCategory\n  (economy/realty/issue/tenancy)\nNewsTier\n  (top3/weekly/tenancy)",
     "kCardNewsList (14편)",
     "N-01, N-02"),
    ("Bill\n(mutable)",
     "models/bill_model.dart",
     "final String id\nBillType type\nString name\nint amount\nint dayOfMonth\nbool autopay\nint get dday\nString get ddayLabel",
     "BillType\n  (rent/manage/electric/gas/\n   internet/mobile/water/\n   insurance/other)",
     "kMockBills (6건)",
     "B-01"),
    ("Offer\n(clicked 가변)",
     "models/offer_model.dart",
     "String id, icon, name, desc\nOfferCategory category\nOfferType type\nint reward\nbool clicked",
     "OfferCategory\n  (rent/telecom/utility/subscribe/\n   manage/insurance/transit/news)\nOfferType\n  (simple/recom/high)",
     "kOffers (29개)",
     "OW-01"),
]

for ri, row in enumerate(models):
    bg = ROW_BG_ALT if ri % 2 == 0 else ROW_BG_NORM
    for ci, val in enumerate(row, 1):
        fg = "1A1A1A"
        if ci == 1: fg = "0D1A3E"
        if ci == 2: fg = "6B7280"
        cell_style(ws2, ri + 3, ci, val, bg=bg, fg=fg, wrap=True, size=9)
    ws2.row_dimensions[ri + 3].height = max(
        15, len(str(row[2]).split("\n")) * 13
    )

ws2.freeze_panes = "A3"


# ════════════════════════════════════════════════
# 시트 3 — AppState 메서드
# ════════════════════════════════════════════════
ws3 = wb.create_sheet("03_AppState메서드")

ws3.merge_cells("A1:E1")
t3 = ws3["A1"]
t3.value = "마이밍 Flutter — AppState 메서드 · 상태 흐름"
t3.fill = fill("1A3A7A")
t3.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
t3.alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 14

headers3 = ["메서드", "시그니처", "동작 설명", "호출 화면", "영속화"]
header_row(ws3, 2, headers3)
ws3.row_dimensions[2].height = 28

col_w3 = [18, 50, 46, 32, 14]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

methods = [
    ("earn()", "void earn(String icon, String name, int amount)",
     "walletAmount += amount\naddLedger() 추가\nnotifyListeners()\nsaveToPrefs() ← 앱 재시작 후 복원",
     "H-03, N-01, N-02\nG-02~G-05, F-01, OW-01", "✅ saveToPrefs()"),
    ("spend()", "void spend(int amount)",
     "walletAmount -= amount (잔액 부족 시 무시)\naddLedger() 추가\nnotifyListeners()\nsaveToPrefs()",
     "SP-01 (이용권/상품권/주거)\nB-01", "✅ saveToPrefs()"),
    ("buyBooster()", "void buyBooster(String effect, String name, int price)",
     "walletAmount -= price\nhasManhwaBoost/hasSpeed2x/hasSpeed5x/hasAutoCollect 플래그 설정\naddLedger() 추가\nnotifyListeners()\nsaveToPrefs()",
     "SP-01 (부스터 탭)", "✅ saveToPrefs()"),
    ("addLedger()", "void addLedger(icon, name, amount, type)",
     "LedgerEntry 생성 후 ledger.insert(0, …)\n최대 100건 유지",
     "earn() / spend() / buyBooster() 내부 자동 호출", "간접 (상위 메서드 통해)"),
    ("readCNArticle()", "void readCNArticle()",
     "cnMissionDone=true면 즉시 return\ncnReadCount++\nwalletAmount += 10, addLedger()\n3회 달성 시 walletAmount += 30, cnMissionDone=true\nnotifyListeners()\nsaveToPrefs()",
     "N-01 (카드 탭 시)", "✅ saveToPrefs()"),
    ("doAttendance()", "void doAttendance()",
     "canAttendToday=false면 즉시 return\nlastAttendanceDate = 오늘\nattendanceDates에 오늘 날짜 추가\nearn(+₩10)",
     "H-03", "✅ (earn 내부)"),
    ("canAttendToday", "bool get canAttendToday",
     "lastAttendanceDate != 오늘 날짜 (ISO 기준)",
     "H-03", "—"),
    ("addSteps()", "void addSteps(int count)",
     "steps += count\nrate = hasManhwaBoost ? 0.015 : 0.01\n(10,000보 = ₩150 부스트 / ₩100 기본)\nwalletAmount += count * rate\naddLedger()\nnotifyListeners()\nsaveToPrefs()",
     "H-01 (pedometer 스트림)", "✅ saveToPrefs()"),
    ("startSession()", "void startSession()",
     "firstAccessDate ??= 오늘 날짜 (최초 1회)\n_sessionStart ??= DateTime.now()\n⚠ totalSecondsInApp은 저장하지 않음\n→ 앱 종료(kill) 시 자동 초기화",
     "main.dart\n(initState + AppLifecycleState.resumed)", "firstAccessDate만 저장"),
    ("pauseSession()", "void pauseSession()",
     "totalSecondsInApp += 경과 초\n_sessionStart = null\n백그라운드 전환 시 호출\n앱 재시작 후 totalSecondsInApp = 0",
     "main.dart\n(AppLifecycleState.paused / dispose)", "저장 안 함 (세션 한정)"),
    ("currentSessionSeconds", "int get currentSessionSeconds",
     "_sessionStart == null → 0\n그 외: DateTime.now().difference(_sessionStart).inSeconds",
     "SP-01 _DwellTimeCard", "—"),
    ("totalSecondsWithCurrent", "int get totalSecondsWithCurrent",
     "totalSecondsInApp + currentSessionSeconds\n앱 재시작 후엔 totalSecondsInApp=0\n→ = currentSessionSeconds",
     "SP-01 _DwellTimeCard", "—"),
    ("switchToCardNewsTab()", "void switchToCardNewsTab()",
     "cardNewsTabTrigger++ → notifyListeners()\nMainShell이 감지해 카드뉴스 탭(index 1)으로 전환",
     "H-03 광고보너스 탭", "—"),
    ("loadFromPrefs()", "Future<void> loadFromPrefs()",
     "SharedPreferences에서 전체 상태 복원\nattendanceDates 로드 (getStringList)\nnotifyListeners()\n⚠ totalSecondsInApp은 로드 안 함\n→ 앱 시작마다 0으로 초기화",
     "main.dart (앱 시작)", "읽기 전용"),
    ("saveToPrefs()", "Future<void> saveToPrefs()",
     "steps, wallet, 부스터 플래그 4종\ncnRead, cnMission, attendance 저장\nattendanceDates (전체 출석 날짜 목록) 저장\nfirstAccessDate 저장\n⚠ totalSecondsInApp은 저장 안 함",
     "earn/spend/buyBooster/addSteps\nstartSession 후 자동 호출", "쓰기"),
]

for ri, row in enumerate(methods):
    bg = ROW_BG_ALT if ri % 2 == 0 else ROW_BG_NORM
    for ci, val in enumerate(row, 1):
        fg = "1A1A1A"
        bold = False
        if ci == 1: fg = "0D1A3E"; bold = True
        if ci == 2: fg = "0D6E6E"
        if ci == 5:
            fg = "00875A" if "✅" in str(val) else "8B95A1"
        cell_style(ws3, ri + 3, ci, val, bg=bg, fg=fg, bold=bold, wrap=True, size=9)
    ws3.row_dimensions[ri + 3].height = max(
        15, len(str(row[2]).split("\n")) * 13
    )

ws3.freeze_panes = "A3"


# ════════════════════════════════════════════════
# 시트 4 — 게임 스펙
# ════════════════════════════════════════════════
ws4 = wb.create_sheet("04_게임스펙")

ws4.merge_cells("A1:H1")
t4 = ws4["A1"]
t4.value = "마이밍 Flutter — 게임 스펙 비교"
t4.fill = fill("1A3A7A")
t4.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
t4.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[1].height = 14

headers4 = ["게임", "ID", "엔진", "조작", "핵심 컴포넌트", "보상 공식", "최대 보상", "Mixin"]
header_row(ws4, 2, headers4)
ws4.row_dimensions[2].height = 28

col_w4 = [16, 8, 14, 26, 34, 32, 10, 40]
for i, w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

games = [
    ("🧱 벽돌깨기", "G-02", "Flame", "화면 탭 (시작) + 패들 드래그",
     "BrickGame extends FlameGame\nBall, Paddle, Brick (8×5)\n_ResultOverlay",
     "Flutter: min(score ÷ 10, 200)\nSPA: 레벨별 earn(lv1:₩15~lv10:₩90)", "Flutter:₩200 / SPA:₩90",
     "HasCollisionDetection\nTapCallbacks\nPanDetector"),
    ("🏃 무한점프", "G-03", "Flame", "화면 탭 (점프)",
     "JumpGame extends FlameGame\nRunner, Obstacle, Coin",
     "min(dist ÷ 20 + coins × 5, 150)", "₩150",
     "TapCallbacks"),
    ("🎣 낚시대전", "G-04", "Flame", "좌우 드래그",
     "CatchGame extends FlameGame\nBucket, Fish, Coin, Bomb\n🐟+10 🪙+15 💣−10",
     "min(max(score,0) ÷ 15, 100)", "₩100",
     "PanDetector\nTapCallbacks"),
    ("🐍 뱀게임", "G-05", "CustomPainter", "방향키 버튼 + 스와이프",
     "_SnakePainter (Canvas 직접 렌더링)\nTimer(150ms)\n_DirBtn ×4",
     "min(score × 3, 150)", "₩150",
     "—"),
    ("🧩 퍼즐", "G-01", "개발예정", "—", "—", "—", "—", "—"),
    ("❓ 경제 퀴즈", "G-01", "개발예정", "—", "—", "—", "—", "—"),
    ("🎯 슈팅게임", "G-01", "개발예정", "—", "—", "—", "—", "—"),
    ("🃏 카드배틀", "G-01", "개발예정", "—", "—", "—", "—", "—"),
]

engine_colors = {"Flame": "FFF3E0", "CustomPainter": "E8F0FF", "개발예정": "F1F5F9"}
engine_fg     = {"Flame": "B45309", "CustomPainter": "2C7FFF", "개발예정": "8B95A1"}

for ri, row in enumerate(games):
    bg = ROW_BG_ALT if ri % 2 == 0 else ROW_BG_NORM
    for ci, val in enumerate(row, 1):
        fg = "1A1A1A"
        bold = False
        row_bg = bg
        if ci == 1: bold = True
        if ci == 3:
            row_bg = engine_colors.get(val, bg)
            fg = engine_fg.get(val, "1A1A1A")
            bold = True
        if ci == 7: fg = "00875A"; bold = True
        cell_style(ws4, ri + 3, ci, val, bg=row_bg, fg=fg, bold=bold,
                   wrap=True, size=9)
    ws4.row_dimensions[ri + 3].height = max(
        15, len(str(row[4]).split("\n")) * 14
    )

ws4.freeze_panes = "A3"


# ════════════════════════════════════════════════
# 시트 5 — 패키지 의존성
# ════════════════════════════════════════════════
ws5 = wb.create_sheet("05_패키지의존성")

ws5.merge_cells("A1:E1")
t5 = ws5["A1"]
t5.value = "마이밍 Flutter — 패키지 의존성"
t5.fill = fill("1A3A7A")
t5.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
t5.alignment = Alignment(horizontal="left", vertical="center")
ws5.row_dimensions[1].height = 14

headers5 = ["패키지", "버전", "용도", "사용 화면", "플랫폼 제약"]
header_row(ws5, 2, headers5)
ws5.row_dimensions[2].height = 28
col_w5 = [22, 14, 36, 24, 46]
for i, w in enumerate(col_w5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

packages = [
    ("provider", "^6.1.1", "전역 상태 관리 (ChangeNotifier)", "전체", "—"),
    ("pedometer", "^4.0.2", "실기기 걸음 수 감지\n(Pedometer.stepCountStream)", "H-01, H-02",
     "iOS: Info.plist\n  NSMotionUsageDescription 필수\nAndroid: AndroidManifest\n  ACTIVITY_RECOGNITION 권한 필수"),
    ("google_fonts", "^6.1.0", "Noto Sans KR 폰트", "전체 (AppTheme)", "—"),
    ("shared_preferences", "^2.2.2",
     "지갑·걸음수·미션 영속화\n(앱 재시작 시 복원)", "AppState", "—"),
    ("flame", "^1.18.0",
     "2D 게임 엔진\n(벽돌깨기·무한점프·낚시대전)", "G-02, G-03, G-04", "—"),
]

for ri, row in enumerate(packages):
    bg = ROW_BG_ALT if ri % 2 == 0 else ROW_BG_NORM
    for ci, val in enumerate(row, 1):
        fg = "1A1A1A"
        bold = False
        if ci == 1: fg = "0D1A3E"; bold = True
        if ci == 2: fg = "7C3AED"
        if ci == 5 and val != "—": fg = "D03030"
        cell_style(ws5, ri + 3, ci, val, bg=bg, fg=fg, bold=bold, wrap=True, size=9)
    ws5.row_dimensions[ri + 3].height = max(
        15, len(str(row[4]).split("\n")) * 13
    )

ws5.freeze_panes = "A3"


# ════════════════════════════════════════════════
# 시트 6 — TODO / 미완성 항목
# ════════════════════════════════════════════════
ws6 = wb.create_sheet("06_TODO")

ws6.merge_cells("A1:E1")
t6 = ws6["A1"]
t6.value = "마이밍 Flutter — 미완성 / TODO 항목"
t6.fill = fill("1A3A7A")
t6.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
t6.alignment = Alignment(horizontal="left", vertical="center")
ws6.row_dimensions[1].height = 14

headers6 = ["우선순위", "항목", "관련 화면", "파일", "비고"]
header_row(ws6, 2, headers6)
ws6.row_dimensions[2].height = 28
col_w6 = [12, 46, 14, 40, 38]
for i, w in enumerate(col_w6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

todos = [
    ("MUST", "iOS Info.plist — NSMotionUsageDescription 추가", "H-01",
     "ios/Runner/Info.plist", "없으면 앱스토어 심사 거절"),
    ("MUST", "Android AndroidManifest — ACTIVITY_RECOGNITION 권한", "H-01",
     "android/app/src/main/AndroidManifest.xml", "Android 10+ 필수"),
    ("HIGH", "소셜 OAuth SDK 실제 연동 (카카오/네이버/구글/애플)", "A-01",
     "pages/auth/login_page.dart", "현재 0.8초 딜레이 후 자동 진입"),
    ("HIGH", "광고 SDK 연동 (광고 보너스 버튼)", "H-03",
     "home/widgets/quick_action_card.dart", "AdMob 또는 자체 광고"),
    ("HIGH", "기사 상세 광고 배너 연동", "N-02",
     "pages/card_news/article_view_page.dart", "카드 뷰어 하단 광고 슬롯"),
    ("MEDIUM", "게임 4개 추가 개발 (개발예정 스낵바)", "G-01",
     "pages/game/game_select_page.dart", "puzzle·quiz·shooting·card"),
    ("MEDIUM", "랭킹 실서버 연동 (현재 목데이터 10건)", "G-06",
     "pages/game/global_rank_page.dart", "Firebase/REST API"),
    ("MEDIUM", "미션 요약 카드 실시간 연동 (현재 하드코딩)", "H-01",
     "pages/home/home_page.dart", "AppState done 여부 반영"),
    ("LOW", "쿠폰 실제 발급 로직", "C-01",
     "pages/coupon/coupon_page.dart", "현재 목데이터만 표시"),
    ("LOW", "프로필 편집 기능 (닉네임·아바타 변경)", "P-01",
     "pages/profile/profile_page.dart", "—"),
]

priority_styles = {
    "MUST":   ("FFE5E5", "D03030"),
    "HIGH":   ("FFF3E0", "E06C00"),
    "MEDIUM": ("EAF4FF", "2C7FFF"),
    "LOW":    ("F1F5F9", "64748B"),
}

for ri, row in enumerate(todos):
    pri = row[0]
    pbg, pfg = priority_styles.get(pri, ("FFFFFF", "1A1A1A"))
    bg = ROW_BG_ALT if ri % 2 == 0 else ROW_BG_NORM
    for ci, val in enumerate(row, 1):
        if ci == 1:
            cell_style(ws6, ri + 3, ci, val, bg=pbg, fg=pfg,
                       bold=True, wrap=False, align_h="center", size=9)
        elif ci == 4:
            cell_style(ws6, ri + 3, ci, val, bg=bg, fg="6B7280",
                       wrap=True, size=8)
        else:
            cell_style(ws6, ri + 3, ci, val, bg=bg, fg="1A1A1A",
                       wrap=True, size=9)
    ws6.row_dimensions[ri + 3].height = 18

ws6.freeze_panes = "A3"


# ── 저장 ─────────────────────────────────────────────────
out = "/Users/freekino_pnwat/Desktop/제작/마이밍 데모 변경/마이밍_flutter/마이밍_IA_문서.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
